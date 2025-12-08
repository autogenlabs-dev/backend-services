"""
Script to upload templates from Excel file to MongoDB database.
"""

import asyncio
from openpyxl import load_workbook
from datetime import datetime, timezone
from motor.motor_asyncio import AsyncIOMotorClient
from bson import ObjectId
import json
import os
from dotenv import load_dotenv

# Load .env file
load_dotenv()

# MongoDB connection - Use DATABASE_URL from .env
MONGODB_URL = os.getenv("DATABASE_URL", "mongodb://localhost:27017/user_management_db")

# Default user ID for templates (you can change this)
DEFAULT_USER_ID = ObjectId("000000000000000000000001")

async def upload_templates():
    """Upload templates from Excel to MongoDB."""
    
    # Connect to MongoDB
    client = AsyncIOMotorClient(MONGODB_URL)
    db = client.get_default_database()  # Get database from connection string
    templates_collection = db.templates
    
    # Load Excel file
    excel_path = "/home/cis/Music/Autogenlabs-Web-App/templates.xlsx"
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # Get headers
    headers = [cell.value for cell in ws[1]]
    print(f"Excel columns: {headers}")
    print(f"Total rows: {ws.max_row - 1}")
    
    # Category mapping to standardize categories
    category_mapping = {
        "3d": "3D & Interactive",
        "landing": "Landing Page",
        "portfolio": "Portfolio",
        "ecommerce": "E-commerce",
        "dashboard": "Dashboard",
        "admin": "Admin Panel",
        "blog": "Blog",
        "saas": "SaaS Tool",
        "agency": "Agency",
    }
    
    # Read all rows and create template documents
    templates = []
    for row_idx in range(2, ws.max_row + 1):
        row_data = {}
        for col_idx, header in enumerate(headers, start=1):
            row_data[header] = ws.cell(row=row_idx, column=col_idx).value
        
        # Parse tags from string to list
        tags = []
        if row_data.get('tags'):
            try:
                tags_str = row_data['tags']
                if isinstance(tags_str, str):
                    tags = json.loads(tags_str.replace("'", '"'))
            except:
                tags = []
        
        # Parse tech_stack from string to list
        tech_stack = []
        if row_data.get('tech_stack'):
            try:
                tech_str = row_data['tech_stack']
                if isinstance(tech_str, str):
                    tech_stack = json.loads(tech_str.replace("'", '"'))
            except:
                tech_stack = []
        
        # Map category
        raw_category = (row_data.get('category') or 'other').lower()
        category = category_mapping.get(raw_category, raw_category.title())
        
        # Infer type/language from tags
        template_type = "React"  # Default
        language = "JavaScript"
        
        if tags:
            tags_lower = [t.lower() for t in tags]
            if 'nextjs' in tags_lower or 'next' in tags_lower:
                template_type = "Next.js"
            elif 'vue' in tags_lower or 'nuxt' in tags_lower:
                template_type = "Vue"
            elif 'angular' in tags_lower:
                template_type = "Angular"
            elif 'svelte' in tags_lower:
                template_type = "Svelte"
            elif 'html' in tags_lower or 'css' in tags_lower:
                template_type = "HTML/CSS"
            
            if 'typescript' in tags_lower:
                language = "TypeScript"
        
        # Create template document
        template = {
            "title": row_data.get('title', 'Untitled'),
            "category": category,
            "type": template_type,
            "language": language,
            "difficulty_level": "Medium",  # Default
            "plan_type": "Free",
            "rating": 0.0,
            "downloads": 0,
            "views": 0,
            "likes": int(row_data.get('stars', 0) or 0),
            "short_description": (row_data.get('description') or '')[:200],
            "full_description": row_data.get('description') or '',
            "preview_images": [row_data.get('thumbnail_url')] if row_data.get('thumbnail_url') else [],
            "git_repo_url": row_data.get('github_url'),
            "live_demo_url": row_data.get('preview_url'),
            "dependencies": tech_stack,
            "tags": tags,
            "developer_name": row_data.get('source', 'Community'),
            "developer_experience": "Unknown",
            "is_available_for_dev": True,
            "featured": False,
            "popular": int(row_data.get('stars', 0) or 0) > 1000,  # Mark popular if > 1000 stars
            "user_id": DEFAULT_USER_ID,
            "code": None,
            "readme_content": None,
            "approval_status": "approved",  # Auto-approve
            "submitted_for_approval_at": datetime.now(timezone.utc),
            "approved_at": datetime.now(timezone.utc),
            "approved_by": DEFAULT_USER_ID,
            "rejection_reason": None,
            "is_purchasable": True,
            "purchase_count": 0,
            "average_rating": 0.0,
            "total_ratings": 0,
            "comments_count": 0,
            "last_comment_at": None,
            "rating_distribution": {
                "5_star": 0,
                "4_star": 0,
                "3_star": 0,
                "2_star": 0,
                "1_star": 0
            },
            "created_at": datetime.now(timezone.utc),
            "updated_at": datetime.now(timezone.utc),
            "is_active": True
        }
        
        templates.append(template)
    
    print(f"\nPrepared {len(templates)} templates for upload")
    
    # Insert templates
    if templates:
        result = await templates_collection.insert_many(templates)
        print(f"✅ Successfully inserted {len(result.inserted_ids)} templates")
        
        # Print first few IDs
        print("\nFirst 5 inserted IDs:")
        for i, _id in enumerate(result.inserted_ids[:5]):
            print(f"  {i+1}. {_id}")
    
    # Close connection
    client.close()
    
    return len(templates)

if __name__ == "__main__":
    count = asyncio.run(upload_templates())
    print(f"\n🎉 Done! Uploaded {count} templates to database")
